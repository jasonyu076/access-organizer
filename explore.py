# ACCESS Organizer
# Turns raw Student ACCESS Scores History Report into an organized
# multi-sheet Excel tracker

import pandas as pd
import os
import re
import sys
import traceback
import tkinter as tk

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference 
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

# SECTION 1 — CONSTANTS

literacy_part_time_cutoff = 3.5    # literacy PL at/above this -> part-time (one-way)
composite_exit_cutoff = 4.8        # literacy PL at/above this -> exited (one-way)
kindergarten_code = "15"           # raw export uses grade "15" for Kindergarten

# raw report's columns in order, as they appear after the year column
raw_columns = [
    "GradeRaw", "B", "C", "D", "E",
    "Listening_SS", "Mode_L", "Speaking_SS", "Mode_S", "Reading_SS", "Mode_R",
    "Writing_SS", "Mode_W", "Mode_Resp_W",
    "Comprehension_SS", "Oral_SS", "Literacy_SS", "Composite_SS",
    "Listening_PL", "Speaking_PL", "Reading_PL", "Writing_PL",
    "Comprehension_PL", "Oral_PL", "Literacy_PL", "Composite_PL",
    "RCDTS"
]

# domains presented on the excel sheet
domains = [
    "Listening", "Speaking", "Reading",
    "Writing", "Oral", "Literacy", "Composite"
]

# values in raw file that mean no score
missing_tokens = {"-", "--", "", "n/a", "na", "none"}

status_order = ["Full-Time", "Remained Part-Time", "Newly Part-Time", "Exited"]
status_fills = {
    "Exited" : "C6EFCE",                # green
    "Newly Part-Time" : "FFEB9C",       # yellow
    "Remained Part-Time" : "F8CBAD",    # orange
    "Full-Time" : "FFC7CE"              # red
}

font = "Times New Roman"
hdr_fill = PatternFill("solid", start_color = "1F4E78")
k_format = '[=0]"K";0'                  # grades stored as 0 displays as "K"

def F(**kwargs):
    return Font(name = font, **kwargs)

thin = Side(style = "thin", color = "BFBFBF")
box = Border(left = thin, right = thin, top = thin, bottom = thin)

# SECTION 2 — READING RAW EXPORTS

def read_raw_export(path):

    warnings = []
    raw = pd.read_excel(path, header = None)

    # searches top rows for "SIS Home School" line
    school = None
    for r in range(min(6, len(raw))):
        for c in range(min(4, raw.shape[1])):
            cell = str(raw.iat[r, c])
            if "SIS Home School" in cell:
                m = re.search(r"SIS Home School:\s*\S+\s+(.+)", cell)
                if m and m.group(1).strip():
                    school = m.group(1).strip()
                else:
                    for c2 in range(c + 1, raw.shape[1]):
                        nxt = raw.iat[r, c2]
                        text = str(nxt).strip() if pd.notna(nxt) else ""
                        if text and "Selection Criteria" not in text and "Sorted by" not in text:
                            school = text
                            break
                break
        if school:
            break
    if not school:
        school = os.path.splitext(os.path.basename(path))[0]
        warnings.append(f"Could not find the school name in the file header; "
                        f"using the file name '{school}' instead.")
        
    # locate each student block
    col0 = raw[0].astype(str)
    block_starts = raw.index[col0.str.contains("Student ID", na = False)].tolist()
    if not block_starts:
        raise ValueError("No 'Student ID' rows found — this does not look like a "
                         "Student ACCESS Scores History Report export.")
    block_bounds = block_starts + [len(raw)]

    records = []
    for i, start in enumerate(block_starts):
        end = block_bounds[i + 1]
        sid = raw.iat[start, 1]
        name = raw.iat[start, 3]
        dob = raw.iat[start, 5]
        if pd.isna(sid) or not str(sid).strip():
            warnings.append(f"Skipped one student block (row {start + 1}): missing Student ID.")
            continue
        sid = str(sid).strip()
        name = str(name).strip() if pd.notna(name) else "(name missing)"
        dob = str(dob).strip() if pd.notna(dob) else ""

        found_year_rows = 0

        for r in range(start + 1, end):
            year = _as_year(raw.iat[r, 0])
            if year is None:
                continue
            row_vals = raw.iloc[r, 1:1 + len(raw_columns)].tolist()
            row_vals += [None] * (len(raw_columns) - len(row_vals))
            records.append([sid, name, dob, year] + row_vals)
            found_year_rows += 1
        if found_year_rows == 0:
            warnings.append(f"Student {name} ({sid}): no test-year rows found.")
    
    df = pd.DataFrame(records, columns = ["StudentID", "Name", "DOB", "Year"] + raw_columns)
    return df, school, warnings

# returns an integer year if cell plausibly holds one
def _as_year(value):
    try:
        year = int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None
    if 2000 <= year <= 2100:
        return year
    else:
        return None
    
# SECTION 3 — CLEANING AND VALIDATION

# normalize grades, drop duplicates
def clean(df, warnings):
    score_cols = [c for c in df.columns if c.endswith("_SS") or c.endswith("_PL")]
    for c in score_cols:
        originally_empty = df[c].isna()
        as_text = df[c].astype(str).str.strip().str.lower()
        df[c] = pd.to_numeric(df[c], errors = "coerce")

        bad = df[c].isna() & ~originally_empty & ~as_text.isin(missing_tokens)

        if bad.any():
            warnings.append(f"{int(bad.sum())} unreadable value(s) in column {c}; "
                            f"treated as blank.")
    
    # Grade: "15" means Kindergarten -> store as 0 so it sorts first and then display as "K"
    g = df["GradeRaw"].astype(str).str.strip().str.lstrip("0")
    g = g.replace({"": "0"})
    g = g.where(df["GradeRaw"].astype(str).str.strip() != kindergarten_code, "0")
    df["GradeNum"] = pd.to_numeric(g, errors = "coerce")
    n_bad_grade = int(df["GradeNum"].isna().sum())

    if n_bad_grade:
        warnings.append(f"{n_bad_grade} row(s) had an unreadable grade; left blank.")
    
    dupes = df.duplicated(["StudentID", "Year"], keep = "last")
    if dupes.any():
        warnings.append(f"Removed {int(dupes.sum())} duplicate student-year row(s) "
                        f"(kept the last occurrence of each).")
        df = df[~dupes]
    
    return df.sort_values(["Name", "StudentID", "Year"]).reset_index(drop = True)

def compute_statuses(df):

    def per_student(group):

        lit_best = comp_best = 0.0
        out = []
        for _, row in group.iterrows():
            lit_before = lit_best
            if pd.notna(row.Literacy_PL):
                lit_best = max(lit_best, row.Literacy_PL)
            if pd.notna(row.Composite_PL):
                comp_best = max(comp_best, row.Composite_PL)
            if comp_best >= composite_exit_cutoff:
                out.append("Exited")
            elif lit_best >= literacy_part_time_cutoff:
                if lit_before >= literacy_part_time_cutoff:
                    out.append("Remained Part-Time")
                else:
                    out.append("Newly Part-Time")
            else:
                out.append("Full-Time")
        
        return pd.Series(out, index = group.index)
    
    df["Status"] = (df.groupby("StudentID", group_keys = False)
                    .apply(per_student, include_groups = False))
    
    return df

# SECTION 4 — CREATING THE WORKBOOK

pl_letters = {
    "Listening" : "O",
    "Speaking" : "P", 
    "Reading" : "Q",
    "Writing" : "R", 
    "Oral" : "T",
    "Literacy" : "U",
    "Composite" : "V"
}

ss_letters = {
    "Listening" : "G",
    "Speaking" : "H", 
    "Reading" : "I",
    "Writing" : "J", 
    "Oral" : "L",
    "Literacy" : "M",
    "Composite" : "N"
}

def build_workbook(df, school, out_path):
    years = sorted(df["Year"].unique())
    latest = df.loc[df.groupby("StudentID")["Year"].idxmax()]
    severity = {s: i for i, s in enumerate(status_order)}
    roster = latest.assign(_sev = latest["Status"].map(severity)).sort_values(["_sev", "Name"])

    n_data = len(df)
    data_last = n_data + 1
    key = f"'Data (All Years)'!$X$2:$X${data_last}"

    def dcol(letter):
        return f"'Data (All Years)'!${letter}$2:${letter}${data_last}"
    
    wb = Workbook()
    _sheet_data(wb, df, data_last)
    _sheet_roster(wb, roster, school, key, dcol)
    _sheet_history(wb, roster, school, years, key, dcol)
    list_last = _sheet_lists(wb, latest)
    _sheet_lookup(wb, years, list_last, key, dcol, wb["Lists"])
    _sheet_readme(wb, school, len(latest), n_data, years)
    _sheet_summary(wb, roster)
    del wb["Sheet"]
    wb.save(out_path)

def _header_cells(ws, row, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row = row, column = c)
            cell.fill = hdr_fill
            cell.font = F(bold = True, color = "FFFFFF", size = 10)
            cell.alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)

def _status_cf(ws, cell_range, anchor):
        for status, color in status_fills.items():
            ws.conditional_formatting.add(cell_range, FormulaRule(
                formula = [f'{anchor} = "{status}"'], 
                fill = PatternFill("solid", start_color = color, end_color = color, fill_type = "solid")))

def _heatmap(ws, cell_range):

    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type = "num", start_value = 1, start_color = "F8696B",
        mid_type = "num", mid_value = literacy_part_time_cutoff, mid_color = "FFEB84",
        end_type = "num", end_value = 6, end_color = "63BE7B"))
    
def _sheet_data(wb, df, data_last):
    ws = wb.create_sheet("Data (All Years)")
    headers = (["School", "Student ID", "Student Name", "Date of Birth", "Test Year", "Grade"]
               + [f"{d} SS" for d in ["Listening", "Speaking", "Reading", "Writing",
                                      "Comprehension", "Oral", "Literacy", "Composite"]]
                + [f"{d} PL" for d in ["Listening", "Speaking", "Reading", "Writing", 
                                       "Comprehension", "Oral", "Literacy", "Composite"]]
                + ["Status", "Key"])
    ws.append(headers)

    ss = [
        "Listening_SS", "Speaking_SS", "Reading_SS", "Writing_SS",
          "Comprehension_SS", "Oral_SS", "Literacy_SS", "Composite_SS"
          ]
    pl = [
        "Listening_PL", "Speaking_PL", "Reading_PL", "Writing_PL",
        "Comprehension_PL", "Oral_PL", "Literacy_PL", "Composite_PL"
    ]

    for i, (_, r) in enumerate(df.iterrows()):
        row = 2 + i
        values = [r.School, r.StudentID, r.Name, r.DOB, int(r.Year),
                  int(r.GradeNum) if pd.notna(r.GradeNum) else None]
        values += [r[c] if pd.notna(r[c]) else None for c in ss]
        values += [r[c] if pd.notna(r[c]) else None for c in pl]
        ws.append(values)

        ws.cell(row = row, column = 23).value = (
            f'=IF(_xlfn.MAXIFS($V$2:$V${data_last},$B$2:$B${data_last},$B{row},'
            f'$E$2:$E${data_last},"<="&$E{row}) >= {composite_exit_cutoff},"Exited",'
            f'IF(_xlfn.MAXIFS($U$2:$U${data_last},$B$2:$B${data_last},$B{row},'
            f'$E$2:$E${data_last},"<="&$E{row}) >= {literacy_part_time_cutoff},'
            f'IF(_xlfn.MAXIFS($U$2:$U${data_last},$B$2:$B${data_last},$B{row},'
            f'$E$2:$E${data_last},"<"&$E{row}) >= {literacy_part_time_cutoff},'
            f'"Remained Part-Time","Newly Part-Time"),"Full-Time"))')
        ws.cell(row = row, column = 24).value = f'=$B{row}&"|"&$E{row}'

    _header_cells(ws, 1, 24)
    for row_cells in ws.iter_rows(min_row = 2, max_row = data_last):
        for cell in row_cells:
            cell.font = F(size = 10)
    for r in range(2, data_last + 1):
        ws.cell(row = r, column = 6).number_format = k_format
        for c in range(15, 23):
            ws.cell(row = r, column = c).number_format = "0.0"
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:W{data_last}"
    for i, w in enumerate([17, 11, 22, 11, 9, 7] + [8] * 16 + [17, 14],  1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["X"].hidden = True
    _status_cf(ws, f"W2:W{data_last}", "$W2")

"""
def _sheet_roster(wb, roster, school, key, dcol):
    ws = wb.create_sheet("Current Roster")
    ws["A1"] = f"{school} — Current EL Roster (each student's most recent ACCESS test)"
    ws["A1"].font = F(bold = True, size = 13)
    ws["A2"] = (
        "Color Legend: Green = Exited (Composite PL ≥ 4.8)  "
        "Yellow = Newly Part-Time (first year Literacy PL ≥ 3.5)    "
        "Orange = Remained Part-Time (qualified in a prior year)    "
        "Red = Full-Time / Needs Attention (never reached Literacy 3.5)"
    )
    ws["A2"].font = F(italic = True, size = 9)
    headers = (
        ["Student ID", "Student Name", "Date of Birth", "Grade", "Latest Test Year"]
        + [f"{d} SS" for d in domains] + [f"{d} PL" for d in domains]
        + ["Prior-Yr Composite PL", "Status", "MatchRow"]
    )
    ws.append([])
    ws.append(headers)
    first = 5

    for i, (_, r) in enumerate(roster.iterrows()):
        row = first + i
        ws.cell(row = row, column = 1).value = str(r.StudentID)
        ws.cell(row = row, column = 22).value = (
            f'=IFERROR(MATCH($A{row}&"|"&_xlfn.MAXIFS({dcol("E")},{dcol("B")},$A{row}),'
            f'{key},0),"")'
        )
        for col, letter in [(2, "C"), (3, "D"), (4, "F"), (5, "E")]:
            ws.cell(row = row, column = col).value = f"=INDEX({dcol(letter)},$V{row})"
        for j, d in enumerate(domains):
            for offset, letters in [(6, ss_letters), (13, pl_letters)]:
                ref = f"INDEX({dcol(letters[d])},$V{row})"
                ws.cell(row = row, column = offset + j).value = \
                    f'=IFERROR(IF({ref} = 0, "", {ref}), "")'
        prior = f'INDEX({dcol("V")}, MATCH($A{row}&"|"&($E{row}-1), {key},0))'
        ws.cell(row = row, column = 20).value = f'=IFERROR(IF({prior} = 0, "—", {prior}), "—")'
        ws.cell(row = row, column = 21).value = f'=IFERROR(INDEX({dcol("W")}, $V{row}), "")'
    last = first + len(roster) - 1
    _header_cells(ws, 4, 22)
    for row_cells in ws.iter_rows(min_row = first, max_row = last):
        for cell in row_cells:
            cell.font = F(size = 10)
            cell.border = box
    for r in range(first, last + 1):
        ws.cell(row = r, column = 4).number_format = k_format
        for c in range(13, 21):
            ws.cell(row = r, column = c).number_format = "0.0"
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:U{last}"
    for i, w in enumerate([11, 22, 11, 7, 8] + [8] * 14 + [17, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["V"].hidden = True
    _status_cf(ws, f"A{first}:U{last}", f"$U{first}")
"""

def _sheet_roster(wb, roster, school, key, dcol):
    ws = wb.create_sheet("Current Roster")
    ws["A1"] = f"{school} — Current EL Roster (each student's most recent ACCESS test)"
    ws["A1"].font = F(bold = True, size = 13)
    ws["A2"] = (
        "Color Legend: Green = Exited (Composite ≥ 4.8)     "
        "Yellow = Newly Part-Time (first year Literacy PL ≥ 3.5)    "
        "Orange = Remained Part-Time (qualified in a prior year)    "
        "Red = Full-Time / Needs Attention (never reached Literacy 3.5) "
    )
    ws["A2"].font = F(italic = True, size = 9)
    headers = (
        ["Student ID", "Student Name", "Trend", "Date of Birth", "Grade",
         "Latest Test Year"]
        + [f"{d} SS" for d in domains] + [f"{d} PL" for d in domains]
        + ["Prior-Yr Composite PL", "Status", "MatchRow"]
    )
    ws.append([])
    ws.append(headers)
    first = 5

    for i, (_, r) in enumerate(roster.iterrows()):
        row = first + i
        ws.cell(row = row, column = 1).value = str(r.StudentID)
        ws.cell(row = row, column = 23).value = (
            f'=IFERROR(MATCH($A{row}&"|"&_xlfn.MAXIFS({dcol("E")},{dcol("B")},$A{row}),'
            f'{key},0),"")'
        )
        for col, letter in [(2, "C"), (4, "D"), (5, "F"), (6, "E")]:
            ws.cell(row = row, column = col).value = f"=INDEX({dcol(letter)},$W{row})"
        for j, d in enumerate(domains):
            for offset, letters in [(7, ss_letters), (14, pl_letters)]:
                ref = f"INDEX({dcol(letters[d])},$W{row})"
                ws.cell(row = row, column = offset + j).value = \
                    f'=IFERROR(IF({ref} = 0, "", {ref}), "")'
        prior = f'INDEX({dcol("V")}, MATCH($A{row}&"|"&($F{row}-1), {key},0))'
        ws.cell(row = row, column = 21).value = f'=IFERROR(IF({prior} = 0, "—", {prior}), "—")'
        ws.cell(row = row, column = 22).value = f'=IFERROR(INDEX({dcol("W")}, $W{row}), "")'
        ws.cell(row = row, column = 3).value = (
            f'=IF(OR($T{row}="",$U{row}="—",$U{row}=""),"",'
            f'IF($T{row}>$U{row},"▲",IF($T{row}<$U{row},"▼","—")))'
        )
    last = first + len(roster) - 1
    _header_cells(ws, 4, 23)
    for row_cells in ws.iter_rows(min_row = first, max_row = last):
        for cell in row_cells:
            cell.font = F(size = 10)
            cell.border = box
    for r in range(first, last + 1):
        ws.cell(row = r, column = 5).number_format = k_format
        for c in range(14, 22):
            ws.cell(row = r, column = c).number_format = "0.0"
        ws.cell(row = r, column = 3).alignment = Alignment(horizontal = "center")
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:V{last}"
    for i, w in enumerate([11, 22, 8, 11, 7, 8] + [8] * 14 + [17, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["W"].hidden = True

    for status, color in status_fills.items():
        ws.conditional_formatting.add(
            f"B{first}:B{last}",
            FormulaRule(formula = [f'$V{first}="{status}"'],
                        fill = PatternFill("solid", start_color = color,
                                           end_color = color, fill_type = "solid"))
        )
        ws.conditional_formatting.add(
            f"V{first}:V{last}",
            FormulaRule(formula = [f'$V{first}="{status}"'],
                        fill = PatternFill("solid", start_color = color,
                                            end_color = color, fill_type = "solid"))
        )

    ws.conditional_formatting.add(
        f"C{first}:C{last}",
        FormulaRule(formula = [f'$C{first}="▲"'],
                    font = Font(name = font, color = "008000", bold = True))
    )
    ws.conditional_formatting.add(
        f"C{first}:C{last}",
        FormulaRule(formula = [f'$C{first}="▼"'],
                    font = Font(name = font, color = "CC0000", bold = True))
    )

def _sheet_history(wb, roster, school, years, key, dcol):
    ws = wb.create_sheet("Domain History")
    ws["A1"] = (f"{school} — Proficiency Levels by Domain, All Years"
                f"(1.0 = Entering ... 6.0 = Reaching)")
    ws["A1"].font = F(bold = True, size = 13)
    n_yr = len(years)
    for c, h in [(1, "Student ID"), (2, "Student Name"), (3, "Current Status")]:
        ws.cell(row = 4, column = c).value = h
    for di, d in enumerate(domains):
        c0 = 4 + di * n_yr
        ws.merge_cells(start_row = 3, start_column = c0, end_row = 3, end_column = c0 + n_yr - 1)
        hc = ws.cell(row = 3, column = c0)
        hc.value = f"{d} PL"
        hc.fill = hdr_fill
        hc.font = F(bold = True, color = "FFFFFF", size = 11)
        hc.alignment = Alignment(horizontal = "center")
        for yi, y in enumerate(years):
            ws.cell(row = 4, column = c0 + yi).value = str(y)
    first = 5
    for i, (_, r) in enumerate(roster.iterrows()):
        row = first + i
        ws.cell(row = row, column = 1).value = str(r.StudentID)
        ws.cell(row = row, column = 2).value = r.Name
        ws.cell(row = row, column = 3).value = (
            f'=IFERROR(INDEX({dcol("W")}, MATCH($A{row}&"|"&'
            f'_xlfn.MAXIFS({dcol("E")}, {dcol("B")}, $A{row}), {key}, 0)), "")'
        )
        for di, d in enumerate(domains):
            for yi, y in enumerate(years):
                ref = f'INDEX({dcol(pl_letters[d])}, MATCH($A{row}&"|"&{y}, {key}, 0))'
                ws.cell(row = row, column = 4 + di * n_yr + yi).value = \
                    f'=IFERROR(IF({ref} = 0, "", {ref}), "")'
    last = first + len(roster) - 1
    max_col = 3 + len(domains) * n_yr
    _header_cells(ws, 4, max_col)
    for row_cells in ws.iter_rows(min_row = first, max_row = last, max_col = max_col):
        for cell in row_cells:
            cell.font = F(size = 9)
            if cell.column >= 4:
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal = "center")
    ws.freeze_panes = "D5"
    for col, w in [("A", 11), ("B", 22), ("C", 16)]:
        ws.column_dimensions[col].width = w
    for c in range(4, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6
    for di in range(len(domains)):
        c0 = 4 + di * n_yr
        _heatmap(ws, f"{get_column_letter(c0)}{first}:"
                f"{get_column_letter(c0 + n_yr - 1)}{last}")
    _status_cf(ws, f"C{first}:C{last}", f"$C{first}")

def _sheet_lists(wb, latest):
    ws = wb.create_sheet("Lists")
    ws["A1"] = "Student Picker"
    for i, (_, r) in enumerate(latest.sort_values("Name").iterrows()):
        ws.cell(row = 2 + i, column = 1).value = f"{r.Name}    ({r.StudentID})"
    ws.sheet_state = "hidden"
    return 1 + len(latest)

def _sheet_lookup(wb, years, list_last, key, dcol, lists_wc):
    ws = wb.create_sheet("Student Lookup")
    ws["A1"] = "Student Lookup — pick a student to see their full ACCESS history"
    ws["A1"].font = F(bold = True, size = 13)
    ws["A3"] = "Students:"
    ws["A3"].font = F(bold = True)
    dv = DataValidation(type = "list", formula1 = f"=Lists!$A$2:$A${list_last}",
                        allow_blank = True)
    ws.add_data_validation(dv)
    ws["B3"] = lists_wc["A2"].value
    dv.add(ws["B3"])
    ws["B3"].fill = PatternFill("solid", start_color = "FFF2CC")
    ws["B3"].font = F(bold = True, size = 11)
    ws.merge_cells("B3:E3")
    ws["G3"] = "Student ID:"
    ws["G3"].font = F(bold = True)
    ws["H3"] = ('=TRIM(MID($B$3,FIND("(",$B$3)+1,'
                'FIND(")",$B$3)-FIND("(",$B$3)-1))')
    ws["G4"] = "Date of Birth:"
    ws["G4"].font = F(bold=True)
    ws["H4"] = f'=IFERROR(INDEX({dcol("D")},MATCH($H$3,{dcol("B")},0)),"")'
    headers = ["Year", "Grade"] + [f"{d} PL" for d in domains] + ["Composite SS", "Status"]
    hrow = 6
    for c, h in enumerate(headers, 1):
        ws.cell(row = hrow, column = c).value = h
    _header_cells(ws, hrow, len(headers))
    for yi, y in enumerate(years):
        row = hrow + 1 + yi
        ws.cell(row = row, column = 1).value = str(y)
        m = f'MATCH($H$3&"|"&{y}, {key}, 0)'
        ws.cell(row = row, column = 2).value = f'=IFERROR(INDEX({dcol("F")},{m}),"")'
        for di, d in enumerate(domains):
            ref = f"INDEX({dcol(pl_letters[d])}, {m})"
            ws.cell(row = row, column = 3 + di).value = f'=IFERROR(IF({ref} = 0, NA(), {ref}), NA())'
        ref = f'INDEX({dcol("N")}, {m})'
        ws.cell(row = row, column = 10).value = f'=IFERROR(IF({ref} = 0, "", {ref}),"")'
        ws.cell(row = row, column = 11).value = f'=IFERROR(INDEX({dcol("W")}, {m}), "")'
    tlast = hrow + len(years)
    for row_cells in ws.iter_rows(min_row = hrow + 1, max_row = tlast, max_col = 11):
        for cell in row_cells:
            cell.font = F(size = 10)
            cell.border = box
            cell.alignment = Alignment(horizontal = "center")
            if 3 <= cell.column <= 9:
                cell.number_format = "0.0"
    for r in range(hrow + 1, tlast + 1):
        ws.cell(row = r, column = 2).number_format = k_format
    _heatmap(ws, f"C{hrow + 1}:I{tlast}")
    _status_cf(ws, f"K{hrow + 1}:K{tlast}", f"$K{hrow + 1}")
    for i, w in enumerate([8, 7] + [11] * 7 + [12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    line = LineChart()
    line.title = "Composite PL Over Time"
    line.y_axis.title = "Composite PL"
    line.x_axis.title = "Year"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 6

    line_data = Reference(ws, min_col = 9, min_row = hrow, max_row = tlast)
    line_cats = Reference(ws, min_col = 1, min_row = hrow + 1, max_row = tlast)
    line.add_data(line_data, titles_from_data = True)
    line.set_categories(line_cats)
    line.legend = None

    s = line.series[0]
    s.smooth = False

    light_blue = "6699CC"

    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill = light_blue, w = 20000)

    s.marker = Marker(symbol = "circle", size = 7)
    s.marker.graphicalProperties = GraphicalProperties()
    s.marker.graphicalProperties.solidFill = light_blue
    s.marker.graphicalProperties.line = LineProperties(solidFill = light_blue)

    line.displayBlanksAs = "gap"
    
    line.y_axis.majorGridlines = None
    line.y_axis.delete = False
    line.x_axis.delete = False

    ws.add_chart(line, "B18")

    line.layout = Layout(
        manualLayout = ManualLayout(
            x = 0.01, y = 0.04,
            w = 0.85, h = 0.68
        )
    )

    ws.conditional_formatting.add(
        f"C{hrow+1}:I{tlast}",
        FormulaRule(formula = [f'ISNA(C{hrow+1})'],
                    font = Font(name = font, color = "FFFFFF"))
    )

def _sheet_readme(wb, school, n_students, n_records, years):
    ws = wb.create_sheet("READ ME", 0)
    lines = [
        ("ACCESS for ELL — Multi-Year Tracking Workbook", True, 14),
        ("", False, 10),
        (f"School in this file: {school}    |   Students: {n_students}  |   ", 
         False, 10),
        (f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
         False, 10),
        ("", False, 10),
        ("HOW THE SHEET WORKS", True, 11),
        ("• Data (All Years) — the master table: one row per student per test year,"
         "all scale scores and proficiency levels. This is the only sheet you paste"
         "new data into.", False, 10),
        ("• Current Roster — one row per student, showing their most recent test, "
        "prior-year composite, and current status. Sorted so students needing the"
        "most attention are on top.", False, 10),
        ("• Domain History — one row per student, proficiency levels in every domain "
        "for every year, color-shaded so growth (or stagnation) is visible at a "
        "glance.", False, 10),
        ("• Student Lookup — pick any student from the dropdown to see their complete "
        "testing history on one screen.", False, 10),
        ("", False, 10),
        ("BUILT-IN RULES", True, 11),
        ('• Grade "15" in the raw export means Kindergarten and displays as "K".',
         False, 10),
        ("• Part-time rule is one-way: the first year a student scores Literacy PL "
        "≥ 3.5 they become Newly Part-Time; every year after that they are Remained "
        "Part-Time, even if score later dips below 3.5.", False, 10),
        ("• Exited: Composite (Overall) PL ≥ 4.8 in any year. Exit status also "
        "persists. Exited students may have no rows in recent years — that is "
        "expected (they stopped testing).", False, 10),
        ("• Full-Time: never reached Literacy 3.5 — these students need the most "
        "attention (red).", False, 10),
        ("• All statuses are live formulas — if you correct a score in the Data "
        "sheet, statuses update automatically.", False, 10),
        ("", False, 10),
        ("COLOR LEGEND", True, 11),
        ("Green = Exited    Yellow = Newly Part-Time    Orange = Remained Part-Time     "
        "Red = Full-Time / Needs Attention", False, 10),
        ("Domain History heat map: red = 1.0 (Entering) → yellow = 3.5 → green = 6.0 "
        "(Reaching).", False, 10),
        ("", False, 10),
        ("THIS FILE WAS GENERATED AUTOMATICALLY", True, 11),
        ("It was produced by the ACCESS Organizer tool from a raw 'Student ACCESS "
         "Scores History Report' export. To rebuild it (e.g. after next year's "
         "testing), run the tool again on a fresh export (provided that the raw "
         "export format remains the same).", False, 10),
    ]
    
    for i, (text, bold, size) in enumerate(lines, 1):
        cell = ws.cell(row = i, column = 1)
        cell.value = text
        cell.font = F(bold = bold, size = size)
        cell.alignment = Alignment(wrap_text = True, vertical = "top")
    ws.column_dimensions["A"].width = 130

def _sheet_summary(wb, roster):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Summary — Status Breakdown"
    ws["A1"].font = F(bold = True, size = 14)

    counts = roster["Status"].value_counts()

    ws["A3"] = "Status"
    ws["B3"] = "Number of Students"
    _header_cells(ws, 3, 2)
    for i, status in enumerate(status_order):
        row = 4 + i
        ws.cell(row = row, column = 1).value = status
        ws.cell(row = row, column = 2).value = int(counts.get(status, 0))
        color = status_fills.get(status)
        if color:
            ws.cell(row = row, column = 1).fill = PatternFill(
                "solid", start_color = color, end_color = color, fill_type = "solid")
    total_row = 4 + len(status_order)
    ws.cell(row = total_row, column = 1).value = "Total"
    ws.cell(row = total_row, column = 1).font = F(bold = True)
    ws.cell(row = total_row, column = 2).value = int(counts.sum())
    ws.cell(row = total_row, column = 2).font = F(bold = True)

    ws.cell(row = 3, column = 4).value = "Grade"
    for j, status in enumerate(status_order):
        ws.cell(row = 3, column = 5 + j).value = status
    _header_cells(ws, 3, 1)          # style "Status by Grade" area is below; header styled next
    for c in range(4, 4 + 1 + len(status_order)):
        cell = ws.cell(row = 3, column = c)
        cell.fill = hdr_fill
        cell.font = F(bold = True, color = "FFFFFF", size = 10)
        cell.alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)

    table = pd.crosstab(roster["GradeNum"], roster["Status"])
    table = table.reindex(columns = status_order, fill_value = 0)
    grades_sorted = sorted(table.index)
    for i, grade in enumerate(grades_sorted):
        row = 4 + i
        label = "K" if grade == 0 else str(int(grade))
        ws.cell(row = row, column = 4).value = label
        for j, status in enumerate(status_order):
            ws.cell(row = row, column = 5 + j).value = int(table.loc[grade, status])

    grade_last = 3 + len(grades_sorted)
    chart_row = max(total_row, grade_last) + 2

    chart = BarChart()
    chart.title = "Students by Status"
    chart.y_axis.title = "Number of Students"
    chart.x_axis.title = "Status"
    chart.y_axis.scaling.max = int(counts.max() * 1.05)
    chart.layout = Layout(
        manualLayout = ManualLayout(
            x = 0.01, y = 0.08,
            w = 0.85, h = 0.70
        )
    )
    data = Reference(ws, min_col = 2, min_row = 3, max_row = 3 + len(status_order))
    cats = Reference(ws, min_col = 1, min_row = 4, max_row = 3 + len(status_order))
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(cats)
    chart.legend = None
    series = chart.series[0]
    for i, status in enumerate(status_order):
        pt = DataPoint(idx = i)
        pt.graphicalProperties.solidFill = status_fills.get(status)
        series.data_points.append(pt)
    ws.add_chart(chart, f"B{chart_row}")

    pie = PieChart()
    pie.title = "Status Distribution"
    pie_data = Reference(ws, min_col = 2, min_row = 3, max_row = 3 + len(status_order))
    pie_cats = Reference(ws, min_col = 1, min_row = 4, max_row = 3 + len(status_order))
    pie.add_data(pie_data, titles_from_data = True)
    pie.set_categories(pie_cats)
    slice_series = pie.series[0]
    for i, status in enumerate(status_order):
        pt = DataPoint(idx = i)
        pt.graphicalProperties.solidFill = status_fills.get(status)
        slice_series.data_points.append(pt)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showSerName = False
    pie.dataLabels.showCatName = False
    pie.dataLabels.showVal = False
    pie.dataLabels.showLegendKey = False
    ws.add_chart(pie, f"K{chart_row}")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 8

    chart.y_axis.majorGridlines = None
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    chart.y_axis.scaling.min = 0
    chart.x_axis.title = None

# SECTION 5 — PROCESSING FILE

def process_file(path):
    df, school, warnings = read_raw_export(path)
    if df.empty:
        raise ValueError("No student test records could be read from this file.")
    df = clean(df, warnings)
    df = compute_statuses(df)
    df["School"] = school

    latest_year = int(df["Year"].max())
    safe_school = re.sub(r"[^A-Za-z0-9]+", "", school) or "School"
    out_path = os.path.join(os.path.dirname(os.path.abspath(path)),
                            f"ACCESS_Tracker_{safe_school}_{latest_year}.xlsx")
    build_workbook(df, school, out_path)

    counts = (df.loc[df.groupby('StudentID')['Year'].idxmax(), 'Status']
              .value_counts())
    summary = [
        f"{school}",
        f"  Students: {df['StudentID'].nunique()}  "
        f"Test records: {len(df)}   Years: {int(df['Year'].min())}–{latest_year}",
        "   Current Status: " + ", ".join(
            f"{s}: {counts.get(s, 0)}" for s in status_order),
        f"  Saved: {os.path.basename(out_path)}"]
    summary += [f"  Note: {w}" for w in warnings]
    return out_path, "\n".join(summary)

# SECTION 6 – WINDOW USER SEES

def run_gui():

    root = tk.Tk()
    root.withdraw()

    paths = filedialog.askopenfilenames(
        title = "Select one or more raw ACCESS export files",
        filetypes = [("Excel files", "*.xlt *.xls *.xlsx"), ("All files", "*.*")])
    if not paths:
        return
    
    results, errors = [], []
    for p in paths:
        try:
            _, summary = process_file(p)
            results.append(summary)
        except Exception as e:
            errors.append(f"{os.path.basename(p)}:\n    {e}")
    
    report = ""

    if results:
        report += "FINISHED\n\n" + "\n\n".join(results)
    if errors:
        report += "\n\nPROBLEMS\n\n" + "\n\n".join(errors)
        report += ("\n\nIf a file failed, check that it is the 'Student ACCESS "
                  "Scores History Report' export, then try again.")
    (messagebox.showinfo if not errors else messagebox.showwarning)(
        "ACCESS Organizer", report.strip())
    
def main():
    if len(sys.argv) > 1:
        for p in sys.argv[1:]:
            try:
                out, summary = process_file(p)
                print(summary, "\n")
            except Exception:
                print(f"FAILED on {p}:\n{traceback.format_exc()}")
    else:
        run_gui()

if __name__ == "__main__":
    main()